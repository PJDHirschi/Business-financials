#!/usr/bin/env python3
# rainbow_ledger.py
# Local Python + Excel finance app with a Tkinter GUI and monthly rollover.
# Now includes: Smart Help panel, Command Palette, shortcuts, CSV import/export QoL.
# Requirements: pip install pandas openpyxl

import os
import sys
import shutil
import csv
import random
from datetime import datetime, timedelta
import tkinter as tk
from tkinter import ttk, messagebox, simpledialog, filedialog

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

APP_TITLE = "Rainbow Ledger — Local Excel Finance"
APP_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_ROOT = os.path.join(APP_DIR, "data")

EXCEL_FILENAME = "company_finance.xlsx"
EXCEL_PATH = None  # set at runtime by bootstrap_month_rotation()

# ---------- Smart Help KB ----------
HELP_CONTENT = [
    {
        "topic": "Dashboard",
        "q": "What is the Dashboard",
        "a": "The Dashboard shows YTD Income, YTD Expenses, YTD Net, and a P&L by Month preview. Use Refresh to update after you add invoices, transactions, or payroll."
    },
    {
        "topic": "Transactions",
        "q": "How to add a transaction",
        "a": "Go to Transactions. Set Date, Type (income or expense), Category, Amount, Description, Party, and Method. Click Add Transaction. It writes to the Transactions sheet."
    },
    {
        "topic": "Transactions",
        "q": "Import bank CSV",
        "a": "Click Import Bank CSV. Map your CSV columns to Date, Amount, Description, Type, Category, Party, and Method. The app will add rows to the Transactions sheet."
    },
    {
        "topic": "Invoices",
        "q": "Create an invoice",
        "a": "Go to Invoices. Set Date and Due Date. Pick Customer, Item, Qty, Rate, and optional Notes. Click Create Invoice. It creates an invoice row and a Sales income transaction."
    },
    {
        "topic": "Invoices",
        "q": "Mark an invoice paid",
        "a": "Enter the Invoice ID in the Mark Paid box then click Mark Paid. Status becomes Paid and a marker row is added so you can trace the payment later."
    },
    {
        "topic": "Payroll",
        "q": "Add employee",
        "a": "In Payroll, fill Name, Type (hourly or salary), Hourly Rate or Salary, and Tax Rate. Click Add Employee. The employee appears in the table and the Run Payroll Employee list."
    },
    {
        "topic": "Payroll",
        "q": "Run payroll",
        "a": "Pick an Employee, enter Hours (ignored for salary), set Date, and click Run. It creates a Payslip and two expense transactions: Wages and Taxes and Licenses."
    },
    {
        "topic": "Reports",
        "q": "Build reports",
        "a": "In Reports, click Build or Refresh. The app writes P&L by Month, YTD Summary, and Category Totals into the Reports sheet. Use Export Reports (CSV) on the Dashboard to share."
    },
    {
        "topic": "Settings",
        "q": "Change company name",
        "a": "In Settings, edit Company Name then click Save. It updates the Settings sheet."
    },
    {
        "topic": "Categories",
        "q": "Add or remove categories",
        "a": "In Settings, use Categories. Add a name and select income or expense. Click Add. To remove, select a row and click Remove."
    },
    {
        "topic": "Excel",
        "q": "Print invoice or save as PDF",
        "a": "Open Workbook from the top bar. In Excel, use File > Print or Save As > PDF. You can format the Invoices sheet or create a print-friendly sheet for styling."
    },
    {
        "topic": "Rollover",
        "q": "Monthly rollover",
        "a": "Files are stored in data/YYYY-MM. On a new month, the app finalizes last month (builds reports and saves a FINAL copy) then starts a clean workbook that carries over master data."
    },
    {
        "topic": "Keyboard",
        "q": "Keyboard shortcuts",
        "a": "Ctrl+K opens the Command Palette. Ctrl+1 Dashboard, Ctrl+2 Transactions, Ctrl+3 Invoices, Ctrl+4 Customers, Ctrl+5 Payroll, Ctrl+6 Reports, Ctrl+7 Settings, Ctrl+R Refresh."
    },
]

# -------------------------------
# Backend helpers (Excel + pandas)
# -------------------------------

DEFAULT_CATEGORIES = [
    # Expenses (simplified Schedule C-ish)
    "Advertising","Car and Truck Expenses","Contract Labor","Depreciation","Employee Benefit Programs",
    "Insurance (Other)","Interest (Mortgage)","Interest (Other)","Legal and Professional Services",
    "Office Expenses","Rent or Lease (Vehicles)","Rent or Lease (Other)","Repairs and Maintenance",
    "Supplies","Taxes and Licenses","Travel","Meals (50%)","Utilities","Wages","Other Expenses",
    # Income
    "Sales","Services","Other Income"
]

def set_excel_path(path):
    global EXCEL_PATH
    EXCEL_PATH = path

def month_key(dt=None):
    return (dt or datetime.today()).strftime("%Y-%m")

def month_dir(key=None):
    k = key or month_key()
    d = os.path.join(DATA_ROOT, k)
    os.makedirs(d, exist_ok=True)
    return d

def excel_path_for(key=None):
    return os.path.join(month_dir(key), EXCEL_FILENAME)

def _autosize(ws):
    for col in ws.columns:
        max_len = 10
        letter = get_column_letter(col[0].column)
        for cell in col:
            v = cell.value
            if v is not None:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[letter].width = max_len + 2

def create_workbook(path, company_name="My Company"):
    wb = Workbook()

    ws = wb.active
    ws.title = "Settings"
    ws.append(["Key","Value"])
    ws.append(["CompanyName", company_name])
    ws.append(["BaseCurrency","USD"])
    _autosize(ws)

    ws = wb.create_sheet("ChartOfAccounts")
    ws.append(["Category","Type"])
    for c in DEFAULT_CATEGORIES:
        t = "income" if c in ["Sales","Services","Other Income"] else "expense"
        ws.append([c, t])
    _autosize(ws)

    ws = wb.create_sheet("Customers")
    ws.append(["CustomerName","Email","Phone","BillingAddress","Notes"])
    _autosize(ws)

    ws = wb.create_sheet("Vendors")
    ws.append(["VendorName","Email","Phone","Address","Notes"])
    _autosize(ws)

    ws = wb.create_sheet("Employees")
    ws.append(["EmployeeName","Type","HourlyRate","Salary","TaxRate","Notes"])
    _autosize(ws)

    ws = wb.create_sheet("Transactions")
    ws.append(["Date","Type","Category","Description","CustomerOrVendor","Amount","PaymentMethod","Reference","LinkedDoc"])
    _autosize(ws)

    ws = wb.create_sheet("Invoices")
    ws.append(["InvoiceID","Date","DueDate","CustomerName","Item","Qty","Rate","Amount","Status","Notes"])
    _autosize(ws)

    ws = wb.create_sheet("Payslips")
    ws.append(["PayslipID","Date","EmployeeName","Hours","Gross","Tax","Net","Notes"])
    _autosize(ws)

    ws = wb.create_sheet("Reports")
    ws.append(["ReportName","AsOf","Notes"])
    _autosize(ws)

    wb.save(path)

def find_last_workbook():
    if not os.path.isdir(DATA_ROOT):
        return None
    keys = [k for k in os.listdir(DATA_ROOT) if os.path.isdir(os.path.join(DATA_ROOT, k))]
    keys = sorted([k for k in keys if len(k) == 7 and k[4] == "-"])  # YYYY-MM
    for k in reversed(keys):
        p = os.path.join(DATA_ROOT, k, EXCEL_FILENAME)
        if os.path.exists(p):
            return k, p
    return None

def get_default_columns_for_sheet(sheet):
    defaults = {
        "Settings": ["Key","Value"],
        "ChartOfAccounts": ["Category","Type"],
        "Customers": ["CustomerName","Email","Phone","BillingAddress","Notes"],
        "Vendors": ["VendorName","Email","Phone","Address","Notes"],
        "Employees": ["EmployeeName","Type","HourlyRate","Salary","TaxRate","Notes"],
    }
    return defaults.get(sheet, [])

def create_new_month_from_previous(prev_path, current_path):
    masters = ["Settings","ChartOfAccounts","Customers","Vendors","Employees"]
    tx_sheets = {
        "Transactions": ["Date","Type","Category","Description","CustomerOrVendor","Amount","PaymentMethod","Reference","LinkedDoc"],
        "Invoices":     ["InvoiceID","Date","DueDate","CustomerName","Item","Qty","Rate","Amount","Status","Notes"],
        "Payslips":     ["PayslipID","Date","EmployeeName","Hours","Gross","Tax","Net","Notes"],
        "Reports":      ["ReportName","AsOf","Notes"],
    }
    with pd.ExcelWriter(current_path, engine="openpyxl") as xw:
        for s in masters:
            try:
                df = pd.read_excel(prev_path, sheet_name=s)
                if df is None or df.empty:
                    df = pd.DataFrame(columns=get_default_columns_for_sheet(s))
            except Exception:
                df = pd.DataFrame(columns=get_default_columns_for_sheet(s))
            df.to_excel(xw, sheet_name=s, index=False)
        for s, cols in tx_sheets.items():
            pd.DataFrame(columns=cols).to_excel(xw, sheet_name=s, index=False)

def archive_prev_month(prev_key, prev_path):
    set_excel_path(prev_path)
    try:
        build_reports()
    except Exception:
        pass
    final_path = os.path.join(DATA_ROOT, prev_key, f"company_finance_FINAL_{prev_key}.xlsx")
    try:
        shutil.copy2(prev_path, final_path)
    except Exception:
        pass

def bootstrap_month_rotation(force_close=None, switch_to=None):
    os.makedirs(DATA_ROOT, exist_ok=True)
    if switch_to:
        # switch to existing month key if exists (or create if not)
        path = excel_path_for(switch_to)
        if not os.path.exists(path):
            last = find_last_workbook()
            if last:
                _, last_path = last
                create_new_month_from_previous(last_path, path)
            else:
                create_workbook(path, "My Company")
        set_excel_path(path)
        return

    cur_key = month_key()
    cur_path = excel_path_for(cur_key)
    last = find_last_workbook()

    if force_close and last:
        last_key, last_path = last
        archive_prev_month(last_key, last_path)

    if last:
        last_key, last_path = last
        if last_key != cur_key:
            archive_prev_month(last_key, last_path)
            if not os.path.exists(cur_path):
                create_new_month_from_previous(last_path, cur_path)

    if not os.path.exists(cur_path):
        create_workbook(cur_path, "My Company")

    set_excel_path(cur_path)

def ensure_workbook():
    if not os.path.exists(EXCEL_PATH):
        company = simpledialog.askstring("Welcome", "Enter your Company Name:", initialvalue="My Company")
        if not company:
            company = "My Company"
        create_workbook(EXCEL_PATH, company_name=company)

def read_sheet(sheet):
    return pd.read_excel(EXCEL_PATH, sheet_name=sheet)

def write_sheet(df, sheet):
    with pd.ExcelWriter(EXCEL_PATH, engine="openpyxl", mode="a", if_sheet_exists="replace") as xw:
        df.to_excel(xw, sheet_name=sheet, index=False)

def get_company_name():
    try:
        s = read_sheet("Settings")
        row = s.loc[s["Key"]=="CompanyName"]
        if not row.empty:
            return str(row.iloc[0]["Value"])
        return "My Company"
    except Exception:
        return "My Company"

def set_company_name(new_name):
    s = read_sheet("Settings")
    if (s["Key"]=="CompanyName").any():
        s.loc[s["Key"]=="CompanyName", "Value"] = new_name
    else:
        s = pd.concat([s, pd.DataFrame([{"Key":"CompanyName","Value":new_name}])], ignore_index=True)
    write_sheet(s, "Settings")

def get_categories_df():
    try:
        return read_sheet("ChartOfAccounts")
    except Exception:
        create_workbook(EXCEL_PATH, get_company_name())
        return read_sheet("ChartOfAccounts")

def get_categories(kind=None):
    df = get_categories_df()
    if kind in ("income","expense"):
        return sorted(df[df["Type"]==kind]["Category"].astype(str).tolist())
    return sorted(df["Category"].astype(str).tolist())

def add_category(name, ctype):
    df = get_categories_df()
    if ((df["Category"] == name) & (df["Type"] == ctype)).any():
        return
    df = pd.concat([df, pd.DataFrame([{"Category":name, "Type":ctype}])], ignore_index=True)
    write_sheet(df, "ChartOfAccounts")

def remove_category(name):
    df = get_categories_df()
    df = df[df["Category"] != name]
    write_sheet(df, "ChartOfAccounts")

def next_id(df, col, prefix):
    if df.empty:
        return f"{prefix}0001"
    s = (
        df[col].dropna().astype(str)
        .str.replace(prefix,"", regex=False)
        .str.extract(r"(\d+)")
        .dropna()
    )
    if s.empty:
        return f"{prefix}0001"
    n = int(s.max()[0]) + 1
    return f"{prefix}{n:04d}"

def add_customer(name, email="", phone="", billing="", notes=""):
    df = read_sheet("Customers")
    new = pd.DataFrame([{"CustomerName":name,"Email":email,"Phone":phone,"BillingAddress":billing,"Notes":notes}])
    df = pd.concat([df, new], ignore_index=True)
    write_sheet(df, "Customers")

def get_customers():
    try:
        return read_sheet("Customers")["CustomerName"].dropna().astype(str).tolist()
    except Exception:
        return []

def add_employee(name, etype="hourly", hourly_rate=0.0, salary=0.0, taxrate=0.1, notes=""):
    df = read_sheet("Employees")
    new = pd.DataFrame([{"EmployeeName":name,"Type":etype,"HourlyRate":hourly_rate,"Salary":salary,"TaxRate":taxrate,"Notes":notes}])
    df = pd.concat([df, new], ignore_index=True)
    write_sheet(df, "Employees")

def get_employees():
    try:
        return read_sheet("Employees")["EmployeeName"].dropna().astype(str).tolist()
    except Exception:
        return []

def add_transaction(date, ttype, category, amount, description="", party="", paymethod="", reference="", linked=""):
    df = read_sheet("Transactions")
    new = pd.DataFrame([{
        "Date": pd.to_datetime(date),
        "Type": ttype,
        "Category": category,
        "Description": description,
        "CustomerOrVendor": party,
        "Amount": float(amount),
        "PaymentMethod": paymethod,
        "Reference": reference,
        "LinkedDoc": linked
    }])
    df = pd.concat([df, new], ignore_index=True)
    write_sheet(df, "Transactions")

def create_invoice(date, due_date, customer, item, qty, rate, notes=""):
    inv_df = read_sheet("Invoices")
    inv_id = next_id(inv_df, "InvoiceID", "INV")
    amount = float(qty) * float(rate)

    new_inv = pd.DataFrame([{
        "InvoiceID":inv_id,"Date":pd.to_datetime(date),"DueDate":pd.to_datetime(due_date),
        "CustomerName":customer,"Item":item,"Qty":float(qty),"Rate":float(rate),
        "Amount":amount,"Status":"Unpaid","Notes":notes
    }])
    inv_df = pd.concat([inv_df, new_inv], ignore_index=True)
    write_sheet(inv_df, "Invoices")

    add_transaction(date, "income", "Sales", amount, f"Invoice {inv_id}: {item} x{qty}", customer, "Invoice", inv_id, inv_id)
    return inv_id, amount

def mark_invoice_paid(invoice_id, date, method="Bank"):
    inv_df = read_sheet("Invoices")
    if not (inv_df["InvoiceID"] == invoice_id).any():
        raise ValueError("Invoice not found")
    inv_df.loc[inv_df["InvoiceID"] == invoice_id, "Status"] = "Paid"
    write_sheet(inv_df, "Invoices")
    add_transaction(date, "income", "Other Income", 0.0, f"Payment received for {invoice_id}", "", method, invoice_id, invoice_id)

def run_payroll(date, employee, hours=0.0):
    emp_df = read_sheet("Employees")
    row = emp_df[emp_df["EmployeeName"]==employee]
    if row.empty:
        raise ValueError("Employee not found")
    r = row.iloc[0]
    etype = str(r["Type"]).strip().lower()
    taxrate = float(r["TaxRate"] or 0.1)
    if etype == "hourly":
        gross = float(hours) * float(r["HourlyRate"] or 0.0)
    else:
        gross = float(r["Salary"] or 0.0) / 26.0
    tax = round(gross * taxrate, 2)
    net = round(gross - tax, 2)

    ps_df = read_sheet("Payslips")
    ps_id = next_id(ps_df, "PayslipID", "PAY")
    new_ps = pd.DataFrame([{
        "PayslipID":ps_id,"Date":pd.to_datetime(date),"EmployeeName":employee,"Hours":float(hours),
        "Gross":gross,"Tax":tax,"Net":net,"Notes":""
    }])
    ps_df = pd.concat([ps_df, new_ps], ignore_index=True)
    write_sheet(ps_df, "Payslips")

    add_transaction(date, "expense", "Wages", gross, f"Payroll gross {ps_id}", employee, "Bank", ps_id, ps_id)
    add_transaction(date, "expense", "Taxes and Licenses", tax, f"Payroll tax {ps_id}", employee, "Bank", ps_id, ps_id)
    return ps_id, gross, tax, net

def build_reports():
    tx = read_sheet("Transactions")
    if tx.empty:
        rep = pd.DataFrame([{"ReportName":"No data yet","AsOf":pd.Timestamp.today(),"Notes":""}])
        write_sheet(rep, "Reports")
        return rep

    tx["Date"] = pd.to_datetime(tx["Date"])
    tx["YearMonth"] = tx["Date"].dt.to_period("M").astype(str)

    income = tx[tx["Type"]=="income"].groupby("YearMonth")["Amount"].sum().rename("Income")
    exp = tx[tx["Type"]=="expense"].groupby("YearMonth")["Amount"].sum().rename("Expenses")
    pnl = pd.concat([income, exp], axis=1).fillna(0.0)
    pnl["NetProfit"] = pnl["Income"] - pnl["Expenses"]
    pnl = pnl.reset_index().rename(columns={"YearMonth":"Period"})

    ytd_income = tx[tx["Type"]=="income"]["Amount"].sum()
    ytd_exp = tx[tx["Type"]=="expense"]["Amount"].sum()
    ytd = pd.DataFrame([
        {"Metric":"YTD Income","Amount":ytd_income},
        {"Metric":"YTD Expenses","Amount":ytd_exp},
        {"Metric":"YTD Net","Amount":ytd_income - ytd_exp}
    ])

    cat_totals = (tx.groupby(["Type","Category"])["Amount"]
                  .sum().reset_index()
                  .sort_values(["Type","Amount"], ascending=[True, False]))

    with pd.ExcelWriter(EXCEL_PATH, engine="openpyxl", mode="a", if_sheet_exists="replace") as xw:
        pd.DataFrame([{"ReportName":"P&L by Month","AsOf":pd.Timestamp.today(),"Notes":""}]).to_excel(
            xw, "Reports", index=False, startrow=0
        )
        pnl.to_excel(xw, "Reports", index=False, startrow=2)

        start2 = 2 + pnl.shape[0] + 2
        pd.DataFrame([{"ReportName":"YTD Summary","AsOf":pd.Timestamp.today(),"Notes":""}]).to_excel(
            xw, "Reports", index=False, startrow=start2
        )
        ytd.to_excel(xw, "Reports", index=False, startrow=start2+2)

        start3 = start2 + 2 + ytd.shape[0] + 2
        pd.DataFrame([{"ReportName":"Category Totals","AsOf":pd.Timestamp.today(),"Notes":""}]).to_excel(
            xw, "Reports", index=False, startrow=start3
        )
        cat_totals.to_excel(xw, "Reports", index=False, startrow=start3+2)

    return pnl

# -------------------------------
# GUI (Tkinter)
# -------------------------------

class FancyButton(ttk.Button):
    def __init__(self, master, **kwargs):
        super().__init__(master, **kwargs)
        self._base_style = kwargs.get("style", "Accent.TButton")
        self._pulse = False
        self.bind("<Enter>", self._on)
        self.bind("<Leave>", self._off)
        self.after(600, self._pulse_tick)

    def _on(self, e): self.configure(style="Hover.TButton")
    def _off(self, e): self.configure(style=self._base_style)
    def enable_pulse(self, on=True): self._pulse = on
    def _pulse_tick(self):
        if self._pulse:
            cur = self.cget("padding")
            self.configure(padding=(8 if cur == "10" else 10))
        self.after(600, self._pulse_tick)

class GradientBanner(tk.Canvas):
    def __init__(self, master, width=900, height=84, **kwargs):
        super().__init__(master, width=width, height=height, highlightthickness=0, **kwargs)
        self.cols = ["#ff6b6b","#feca57","#48dbfb","#1dd1a1","#5f27cd"]
        self.offset = 0
        self.animate()

    def draw_gradient(self):
        w = self.winfo_width()
        h = self.winfo_height()
        self.delete("grad")
        steps = 40
        for i in range(steps):
            c = self.cols[(i + self.offset) % len(self.cols)]
            x0 = int((i/steps)*w)
            x1 = int(((i+1)/steps)*w)
            self.create_rectangle(x0, 0, x1, h, fill=c, outline="", tags="grad")
        self.delete("title")
        self.create_text(w//2, h//2, text=APP_TITLE, fill="white",
                         font=("Segoe UI", 20, "bold"), tags="title")

    def animate(self):
        self.draw_gradient()
        self.offset = (self.offset + 1) % len(self.cols)
        self.after(120, self.animate)

class ToolTip:
    def __init__(self, widget, text):
        self.widget = widget
        self.text = text
        self.tip = None
        widget.bind('<Enter>', self.show)
        widget.bind('<Leave>', self.hide)
    def show(self, *_):
        if self.tip: return
        x = self.widget.winfo_rootx() + 10
        y = self.widget.winfo_rooty() + self.widget.winfo_height() + 6
        self.tip = tk.Toplevel(self.widget)
        self.tip.overrideredirect(True)
        self.tip.geometry(f"+{x}+{y}")
        lbl = tk.Label(self.tip, text=self.text, bg="#222", fg="white", padx=8, pady=4, font=("Segoe UI", 9))
        lbl.pack()
    def hide(self, *_):
        if self.tip:
            self.tip.destroy()
            self.tip = None

class CommandPalette(tk.Toplevel):
    """Simple command palette with fuzzy-ish filtering."""
    def __init__(self, master, commands):
        super().__init__(master)
        self.title("Command Palette")
        self.geometry("520x360+%d+%d" % (master.winfo_rootx()+80, master.winfo_rooty()+120))
        self.transient(master)
        self.grab_set()
        self.commands = commands  # list of (label, callback)
        self.filtered = list(self.commands)

        self.entry = ttk.Entry(self)
        self.entry.pack(fill="x", padx=10, pady=10)
        self.entry.focus_set()
        self.entry.bind("<KeyRelease>", self._on_change)
        self.entry.bind("<Return>", self._run_selected)
        self.entry.bind("<Escape>", lambda e: self.destroy())

        self.listbox = tk.Listbox(self, height=12)
        self.listbox.pack(fill="both", expand=True, padx=10, pady=(0,10))
        self.listbox.bind("<Double-Button-1>", self._run_selected)
        self._refresh()

    def _score(self, text, query):
        t = text.lower()
        q = query.lower().strip()
        if not q:
            return 1
        score = 0
        for part in q.split():
            if part in t:
                score += 1
        return score

    def _on_change(self, _):
        q = self.entry.get()
        scored = [(self._score(lbl, q), lbl, cb) for (lbl, cb) in self.commands]
        scored.sort(key=lambda x: (-x[0], x[1]))
        self.filtered = [(lbl, cb) for s,lbl,cb in scored if s > 0 or q == ""]
        self._refresh()

    def _refresh(self):
        self.listbox.delete(0, tk.END)
        for lbl, _ in self.filtered[:200]:
            self.listbox.insert(tk.END, lbl)
        if self.filtered:
            self.listbox.selection_set(0)

    def _run_selected(self, _=None):
        if not self.filtered:
            return
        idx = self.listbox.curselection()
        if not idx:
            idx = (0,)
        lbl, cb = self.filtered[idx[0]]
        try:
            self.destroy()
            cb()
        except Exception as e:
            messagebox.showerror("Command failed", str(e))

class RainbowLedgerApp(tk.Tk):
    def __init__(self):
        super().__init__()
        bootstrap_month_rotation()
        ensure_workbook()

        self.title(APP_TITLE)
        self.geometry("1120x760")
        self.minsize(1020, 680)

        self.style = ttk.Style(self)
        try:
            if sys.platform.startswith("win"):
                self.style.theme_use("winnative")
            else:
                self.style.theme_use("clam")
        except Exception:
            pass

        self.style.configure("Accent.TButton", font=("Segoe UI", 10, "bold"), padding=10)
        self.style.configure("Hover.TButton", background="#ffffff")
        self.style.configure("Card.TFrame", relief="flat", borderwidth=0)

        # Banner
        self.banner = GradientBanner(self, height=84)
        self.banner.pack(fill="x")

        # Top utility bar
        util = ttk.Frame(self); util.pack(fill="x", padx=12, pady=(8,0))

        btn_refresh = FancyButton(util, text="Refresh", command=self._refresh_all); btn_refresh.enable_pulse(True); btn_refresh.pack(side="left", padx=4)
        btn_open_file = ttk.Button(util, text="Open Workbook", command=self._open_workbook); btn_open_file.pack(side="left", padx=4)
        ToolTip(btn_open_file, "Open the current month's Excel file")

        btn_open_folder = ttk.Button(util, text="Open Month Folder", command=self._open_month_folder); btn_open_folder.pack(side="left", padx=4)
        ToolTip(btn_open_folder, "Show the folder for this month (data/YYYY-MM)")

        btn_switch = ttk.Button(util, text="Switch Month…", command=self._switch_month_dialog); btn_switch.pack(side="left", padx=4)
        ToolTip(btn_switch, "View or work in a different month")

        btn_close_month = ttk.Button(util, text="Close This Month", command=self._close_month); btn_close_month.pack(side="left", padx=4)
        ToolTip(btn_close_month, "Finalize current month: build reports and save FINAL copy")

        btn_backup = ttk.Button(util, text="Backup Workbook", command=self._backup_workbook); btn_backup.pack(side="left", padx=4)
        ToolTip(btn_backup, "Create a time-stamped backup copy of the current workbook")

        # Notebook
        self.nb = ttk.Notebook(self); self.nb.pack(fill="both", expand=True, padx=12, pady=12)

        self._build_dashboard_tab()
        self._build_transactions_tab()
        self._build_invoices_tab()
        self._build_customers_tab()
        self._build_payroll_tab()
        self._build_reports_tab()
        self._build_settings_tab()
        self._build_help_tab()  # new help tab

        # Shortcuts + palette
        self.bind_all("<Control-k>", lambda e: self._open_command_palette())
        self.bind_all("<Control-K>", lambda e: self._open_command_palette())
        self.bind_all("<Control-r>", lambda e: self._refresh_all())
        self.bind_all("<Control-1>", lambda e: self.nb.select(self.nb.tabs()[0]))
        self.bind_all("<Control-2>", lambda e: self.nb.select(self.nb.tabs()[1]))
        self.bind_all("<Control-3>", lambda e: self.nb.select(self.nb.tabs()[2]))
        self.bind_all("<Control-4>", lambda e: self.nb.select(self.nb.tabs()[3]))
        self.bind_all("<Control-5>", lambda e: self.nb.select(self.nb.tabs()[4]))
        self.bind_all("<Control-6>", lambda e: self.nb.select(self.nb.tabs()[5]))
        self.bind_all("<Control-7>", lambda e: self.nb.select(self.nb.tabs()[6]))
        self.bind_all("<Control-8>", lambda e: self.nb.select(self.nb.tabs()[7]) if len(self.nb.tabs()) > 7 else None)

        self._register_commands()

        self.after(600, lambda: self._toast(f"Welcome, {get_company_name()}!"))
        self.after(1400, lambda: self._hint("Tip: Press Ctrl+K to run commands fast"))

    # Utilities
    def _open_workbook(self):
        path = EXCEL_PATH
        if sys.platform.startswith("win"):
            os.startfile(path)
        elif sys.platform == "darwin":
            os.system(f'open "{path}"')
        else:
            os.system(f'xdg-open "{path}"')

    def _open_month_folder(self):
        path = os.path.dirname(EXCEL_PATH)
        if sys.platform.startswith("win"):
            os.startfile(path)
        elif sys.platform == "darwin":
            os.system(f'open "{path}"')
        else:
            os.system(f'xdg-open "{path}"')

    def _switch_month_dialog(self):
        months = []
        if os.path.isdir(DATA_ROOT):
            for k in sorted(os.listdir(DATA_ROOT)):
                if len(k)==7 and k[4]=='-' and os.path.exists(excel_path_for(k)):
                    months.append(k)
        default = month_key()
        m = simpledialog.askstring("Switch Month", f"Enter month key (YYYY-MM)\nExisting: {', '.join(months) or 'none'}",
                                   initialvalue=default)
        if not m: return
        try:
            bootstrap_month_rotation(switch_to=m)
            ensure_workbook()
            self._toast(f"Switched to {m}")
            self._refresh_all()
        except Exception as e:
            messagebox.showerror("Error", f"Could not switch:\n{e}")

    def _close_month(self):
        try:
            k = month_key()
            bootstrap_month_rotation(force_close=True)
            self._toast(f"{k} finalized.")
        except Exception as e:
            messagebox.showerror("Error", f"Close month failed:\n{e}")

    def _backup_workbook(self):
        try:
            dst = os.path.join(os.path.dirname(EXCEL_PATH), f"backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
            shutil.copy2(EXCEL_PATH, dst)
            self._toast("Backup created")
        except Exception as e:
            messagebox.showerror("Error", f"Backup failed:\n{e}")

    def _refresh_all(self):
        self._load_dashboard()
        self._refresh_tx_table()
        self._refresh_inv_table()
        self._refresh_pay_table()
        self._load_report_preview()
        self._hint("Tip: Press Ctrl+K to run commands fast")

    def _toast(self, msg):
        top = tk.Toplevel(self)
        top.overrideredirect(1); top.attributes("-topmost", True)
        x = self.winfo_x() + self.winfo_width()//2 - 160
        y = self.winfo_y() + 120
        top.geometry(f"320x40+{x}+{y}")
        lbl = tk.Label(top, text=msg, bg="#222", fg="white", font=("Segoe UI", 10))
        lbl.pack(fill="both", expand=True)
        top.after(1700, top.destroy)

    def _hint(self, msg):
        x = self.banner.winfo_width()//2
        tag = "hint"
        self.banner.delete(tag)
        self.banner.create_text(x, 16, text=msg, fill="white", font=("Segoe UI", 10, "italic"), tags=tag)
        self.after(4000, lambda: self.banner.delete(tag))

    def _confetti(self):
        particles = []
        for _ in range(24):
            x = random.randint(30, self.banner.winfo_width()-30)
            y = random.randint(10, self.banner.winfo_height()-10)
            r = random.randint(3,6)
            c = random.choice(["#ffffff","#222222","#ffd166","#06d6a0","#118ab2","#ef476f"])
            p = self.banner.create_oval(x-r, y-r, x+r, y+r, fill=c, outline="")
            particles.append(p)
        def fall():
            for p in particles:
                self.banner.move(p, 0, 3)
            self.after(28, cleanup)
        def cleanup():
            if particles and random.random() > 0.08:
                fall()
            else:
                for p in particles:
                    self.banner.delete(p)
        fall()

    def _card(self, parent):
        f = ttk.Frame(parent, style="Card.TFrame"); f.configure(padding=12); f.pack_propagate(False); return f

    # ----- Dashboard -----
    def _build_dashboard_tab(self):
        tab = ttk.Frame(self.nb); self.nb.add(tab, text="Dashboard")

        top = ttk.Frame(tab); top.pack(fill="x", pady=(8, 10))
        refresh_btn = FancyButton(top, text="Refresh Dashboard", command=self._load_dashboard); refresh_btn.enable_pulse(True); refresh_btn.pack(side="left", padx=4)
        btn_export = ttk.Button(top, text="Export Reports (CSV)", command=self._export_reports_csv); btn_export.pack(side="left", padx=4)
        ToolTip(btn_export, "Exports P&L by Month, YTD, and Category Totals to CSV in this month folder")

        # Onboarding checklist
        ob = self._card(tab); ob.pack(fill="x", padx=6, pady=6)
        ttk.Label(ob, text="Get Started Checklist", font=("Segoe UI", 11, "bold")).pack(anchor="w")
        steps = [
            "1) Add your company name in Settings.",
            "2) Add at least one customer.",
            "3) Create your first invoice or add a transaction.",
            "4) Add an employee if you run payroll.",
            "5) Click Build/Refresh Reports on the Reports tab.",
        ]
        for s in steps:
            ttk.Label(ob, text=s).pack(anchor="w")

        self.stats_frame = ttk.Frame(tab); self.stats_frame.pack(fill="both", expand=True)
        self._load_dashboard()

    def _export_reports_csv(self):
        try:
            pnl = build_reports()
            # Rebuild tables here and export
            tx = read_sheet("Transactions")
            tx["Date"] = pd.to_datetime(tx["Date"])
            tx["YearMonth"] = tx["Date"].dt.to_period("M").astype(str)
            income = tx[tx["Type"]=="income"].groupby("YearMonth")["Amount"].sum().rename("Income")
            exp = tx[tx["Type"]=="expense"].groupby("YearMonth")["Amount"].sum().rename("Expenses")
            pnl = pd.concat([income, exp], axis=1).fillna(0.0)
            pnl["NetProfit"] = pnl["Income"] - pnl["Expenses"]
            pnl = pnl.reset_index().rename(columns={"YearMonth":"Period"})

            ytd_income = tx[tx["Type"]=="income"]["Amount"].sum()
            ytd_exp = tx[tx["Type"]=="expense"]["Amount"].sum()
            ytd = pd.DataFrame([
                {"Metric":"YTD Income","Amount":ytd_income},
                {"Metric":"YTD Expenses","Amount":ytd_exp},
                {"Metric":"YTD Net","Amount":ytd_income - ytd_exp}
            ])
            cat_totals = (tx.groupby(["Type","Category"])["Amount"]
                          .sum().reset_index()
                          .sort_values(["Type","Amount"], ascending=[True, False]))

            outdir = os.path.dirname(EXCEL_PATH)
            pnl.to_csv(os.path.join(outdir, "P&L_by_Month.csv"), index=False)
            ytd.to_csv(os.path.join(outdir, "YTD_Summary.csv"), index=False)
            cat_totals.to_csv(os.path.join(outdir, "Category_Totals.csv"), index=False)
            self._toast("Reports exported as CSV")
            self._confetti()
        except Exception as e:
            messagebox.showerror("Error", f"Export failed:\n{e}")

    def _load_dashboard(self):
        for w in self.stats_frame.winfo_children():
            w.destroy()
        try:
            tx = read_sheet("Transactions")
            if tx.empty:
                ttk.Label(self.stats_frame, text="No data yet. Add transactions or invoices to get started.",
                          font=("Segoe UI", 12)).pack(pady=20)
                return
            tx["Date"] = pd.to_datetime(tx["Date"])
            income = tx[tx["Type"]=="income"]["Amount"].sum()
            exp = tx[tx["Type"]=="expense"]["Amount"].sum()
            net = income - exp

            row = ttk.Frame(self.stats_frame); row.pack(fill="x", pady=6)
            for title, val in [("YTD Income", income), ("YTD Expenses", exp), ("YTD Net", net)]:
                c = self._card(row); c.pack(side="left", expand=True, fill="x", padx=6)
                ttk.Label(c, text=title, font=("Segoe UI", 11, "bold")).pack(anchor="w")
                ttk.Label(c, text=f"${val:,.2f}", font=("Segoe UI", 16)).pack(anchor="w", pady=(6,0))

            tx["YearMonth"] = tx["Date"].dt.to_period("M").astype(str)
            income_m = tx[tx["Type"]=="income"].groupby("YearMonth")["Amount"].sum()
            exp_m = tx[tx["Type"]=="expense"].groupby("YearMonth")["Amount"].sum()
            pnl = pd.concat([income_m, exp_m], axis=1).fillna(0.0)
            pnl["Net"] = pnl.iloc[:,0] - pnl.iloc[:,1]
            pnl = pnl.reset_index().rename(columns={"YearMonth":"Period", pnl.columns[1]:"Income", pnl.columns[2]:"Expenses"})

            frame_tbl = self._card(self.stats_frame); frame_tbl.pack(fill="both", expand=True, padx=6, pady=6)
            ttk.Label(frame_tbl, text="P&L by Month", font=("Segoe UI", 11, "bold")).pack(anchor="w")
            tv = ttk.Treeview(frame_tbl, columns=("Period","Income","Expenses","Net"), show="headings", height=12)
            for col, w in [("Period",120),("Income",160),("Expenses",160),("Net",160)]:
                tv.heading(col, text=col)
                tv.column(col, width=w, anchor="center" if col=="Period" else "e")
            tv.pack(fill="both", expand=True, pady=(6,0))
            show = pnl.tail(12)
            for _, r in show.iterrows():
                tv.insert("", "end", values=(r["Period"], f"${r['Income']:,.2f}", f"${r['Expenses']:,.2f}", f"${r['Net']:,.2f}"))
        except Exception as e:
            ttk.Label(self.stats_frame, text=f"Error loading dashboard: {e}").pack(pady=20)

    # ----- Transactions -----
    def _build_transactions_tab(self):
        tab = ttk.Frame(self.nb); self.nb.add(tab, text="Transactions")

        form = self._card(tab); form.pack(fill="x", padx=6, pady=6)

        row1 = ttk.Frame(form); row1.pack(fill="x", pady=4)
        ttk.Label(row1, text="Date (YYYY-MM-DD)").pack(side="left", padx=(0,6))
        self.tx_date = ttk.Entry(row1); self.tx_date.insert(0, datetime.today().strftime("%Y-%m-%d")); self.tx_date.pack(side="left", padx=6)
        btn_today = ttk.Button(row1, text="Today", command=lambda: self.tx_date.delete(0, tk.END) or self.tx_date.insert(0, datetime.today().strftime("%Y-%m-%d")))
        btn_today.pack(side="left", padx=6)

        ttk.Label(row1, text="Type").pack(side="left", padx=(18,6))
        self.tx_type = ttk.Combobox(row1, values=["income","expense","transfer"], state="readonly", width=12)
        self.tx_type.current(1); self.tx_type.pack(side="left", padx=6)

        ttk.Label(row1, text="Category").pack(side="left", padx=(18,6))
        self.tx_cat = ttk.Combobox(row1, values=get_categories(), state="readonly", width=28); self.tx_cat.pack(side="left", padx=6)

        row2 = ttk.Frame(form); row2.pack(fill="x", pady=4)
        ttk.Label(row2, text="Amount").pack(side="left", padx=(0,6))
        self.tx_amt = ttk.Entry(row2, width=12); self.tx_amt.pack(side="left", padx=6)

        ttk.Label(row2, text="Description").pack(side="left", padx=(18,6))
        self.tx_desc = ttk.Entry(row2, width=40); self.tx_desc.pack(side="left", padx=6)

        row3 = ttk.Frame(form); row3.pack(fill="x", pady=4)
        ttk.Label(row3, text="Party").pack(side="left", padx=(0,6))
        self.tx_party = ttk.Entry(row3, width=20); self.tx_party.pack(side="left", padx=6)

        ttk.Label(row3, text="Payment Method").pack(side="left", padx=(18,6))
        self.tx_method = ttk.Entry(row3, width=20); self.tx_method.pack(side="left", padx=6)

        add_btn = FancyButton(form, text="Add Transaction", command=self._add_tx); add_btn.enable_pulse(True); add_btn.pack(pady=8)

        # Import/Export row
        row4 = ttk.Frame(form); row4.pack(fill="x", pady=4)
        btn_import = ttk.Button(row4, text="Import Bank CSV…", command=self._import_csv_wizard); btn_import.pack(side="left", padx=4)
        ToolTip(btn_import, "Map your CSV columns to Date/Amount/Description/Type/Category/Party/Method")
        btn_export_tx = ttk.Button(row4, text="Export Transactions CSV", command=self._export_transactions_csv); btn_export_tx.pack(side="left", padx=4)
        
        self.tx_table = ttk.Treeview(tab, columns=("Date","Type","Category","Description","Party","Amount","Method","Reference","Linked"), show="headings", height=18)
        for col, w in [
            ("Date",100),("Type",80),("Category",160),("Description",240),
            ("Party",160),("Amount",110),("Method",120),("Reference",120),("Linked",120)
        ]:
            self.tx_table.heading(col, text=col)
            self.tx_table.column(col, width=w, anchor="w" if col in ("Description","Category","Party","Method") else ("e" if col=="Amount" else "center"))
        self.tx_table.pack(fill="both", expand=True, padx=6, pady=(0,6))

        self._refresh_tx_table()

    def _add_tx(self):
        try:
            date = self.tx_date.get().strip()
            ttype = self.tx_type.get().strip()
            cat = self.tx_cat.get().strip()
            amt_text = self.tx_amt.get().strip()
            desc = self.tx_desc.get().strip()
            party = self.tx_party.get().strip()
            method = self.tx_method.get().strip()
            if not date or not ttype or not cat or not amt_text:
                messagebox.showwarning("Missing data", "Date, Type, Category, and Amount are required.")
                return
            amt = float(amt_text)
            # If user chose transfer, still record it (neutral category)
            add_transaction(date, ttype, cat, amt, desc, party, method)
            self._toast("Transaction added")
            self._refresh_tx_table()
            self._load_dashboard()
        except Exception as e:
            messagebox.showerror("Error", f"Could not add transaction:\n{e}")

    def _refresh_tx_table(self):
        for i in self.tx_table.get_children():
            self.tx_table.delete(i)
        try:
            df = read_sheet("Transactions")
            if df.empty:
                return
            df = df.copy()
            df["Date"] = pd.to_datetime(df["Date"]).dt.strftime("%Y-%m-%d")
            for _, r in df.iterrows():
                self.tx_table.insert("", "end", values=(
                    r.get("Date",""), r.get("Type",""), r.get("Category",""),
                    r.get("Description",""), r.get("CustomerOrVendor",""),
                    f"${float(r.get('Amount',0.0)):,.2f}",
                    r.get("PaymentMethod",""), r.get("Reference",""), r.get("LinkedDoc","")
                ))
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load transactions:\n{e}")

    def _export_transactions_csv(self):
        try:
            df = read_sheet("Transactions")
            out = os.path.join(os.path.dirname(EXCEL_PATH), "Transactions.csv")
            df.to_csv(out, index=False)
            self._toast("Transactions exported")
        except Exception as e:
            messagebox.showerror("Error", f"Export failed:\n{e}")

    # --- CSV Import Wizard (simple mapper) ---
    def _import_csv_wizard(self):
        path = filedialog.askopenfilename(
            title="Select Bank CSV",
            filetypes=[("CSV files","*.csv"),("All files","*.*")]
        )
        if not path:
            return
        try:
            peek = pd.read_csv(path, nrows=250)
        except Exception as e:
            messagebox.showerror("Error", f"Could not read CSV:\n{e}")
            return

        cols = list(peek.columns.astype(str))
        fields = [
            ("Date","Date column"),
            ("Amount","Amount column"),
            ("Description","Description column"),
            ("Type","Type column (income/expense) — optional"),
            ("Category","Category column — optional"),
            ("Party","Party column — optional"),
            ("Method","Payment method column — optional")
        ]

        win = tk.Toplevel(self)
        win.title("Import Bank CSV — Map Columns")
        win.geometry("560x420+%d+%d" % (self.winfo_rootx()+120, self.winfo_rooty()+120))
        ttk.Label(win, text=os.path.basename(path), font=("Segoe UI", 10, "bold")).pack(pady=(10,4))
        frm = ttk.Frame(win); frm.pack(fill="both", expand=True, padx=10, pady=10)

        mappings = {}
        for i,(key,label) in enumerate(fields):
            row = ttk.Frame(frm); row.pack(fill="x", pady=4)
            ttk.Label(row, text=label, width=34).pack(side="left")
            cb = ttk.Combobox(row, values=cols, state="readonly")
            # preselect guesses
            guess = None
            low = [c.lower() for c in cols]
            if key=="Date":
                for cand in ("date","posted","post date","transaction date"):
                    if cand in low: guess = cols[low.index(cand)]; break
            elif key=="Amount":
                for cand in ("amount","amt","debit/credit","value"):
                    if cand in low: guess = cols[low.index(cand)]; break
            elif key=="Description":
                for cand in ("description","memo","details","narrative"):
                    if cand in low: guess = cols[low.index(cand)]; break
            elif key=="Type":
                for cand in ("type","dr/cr","credit/debit","direction"):
                    if cand in low: guess = cols[low.index(cand)]; break
            elif key=="Category":
                if "category" in low: guess = cols[low.index("category")]
            elif key=="Party":
                for cand in ("name","payee","party","merchant","customer"):
                    if cand in low: guess = cols[low.index(cand)]; break
            elif key=="Method":
                for cand in ("method","payment method","channel","card"):
                    if cand in low: guess = cols[low.index(cand)]; break
            if guess: cb.set(guess)
            cb.pack(side="left", padx=6, fill="x", expand=True)
            mappings[key] = cb

        # Default category & type if missing
        rowd = ttk.Frame(frm); rowd.pack(fill="x", pady=(14,6))
        ttk.Label(rowd, text="Default Type when missing:", width=34).pack(side="left")
        def_type = ttk.Combobox(rowd, values=["income","expense"], state="readonly", width=12); def_type.current(1); def_type.pack(side="left", padx=6)
        rowc = ttk.Frame(frm); rowc.pack(fill="x", pady=(4,6))
        ttk.Label(rowc, text="Default Category when missing:", width=34).pack(side="left")
        def_cat = ttk.Combobox(rowc, values=get_categories(), state="readonly", width=28); 
        # pick a common expense default
        if "Other Expenses" in get_categories():
            def_cat.set("Other Expenses")
        def_cat.pack(side="left", padx=6)

        btns = ttk.Frame(win); btns.pack(pady=10)
        ttk.Button(btns, text="Cancel", command=win.destroy).pack(side="right", padx=6)
        def do_import():
            req = ("Date","Amount","Description")
            for r in req:
                if not mappings[r].get():
                    messagebox.showwarning("Map required", f"Please map {r}.")
                    return
            try:
                df = read_sheet("Transactions")
                incoming = pd.read_csv(path)
                # Build normalized frame
                def getcol(k):
                    sel = mappings[k].get()
                    return incoming[sel] if sel in incoming.columns else pd.Series([None]*len(incoming))
                nd = pd.DataFrame({
                    "Date": pd.to_datetime(getcol("Date"), errors="coerce"),
                    "Amount": pd.to_numeric(getcol("Amount"), errors="coerce"),
                    "Description": getcol("Description").astype(str),
                    "Type": (getcol("Type").astype(str).str.lower().str.strip()
                             if mappings["Type"].get() else pd.Series([None]*len(incoming))),
                    "Category": getcol("Category").astype(str) if mappings["Category"].get() else None,
                    "CustomerOrVendor": getcol("Party").astype(str) if mappings["Party"].get() else None,
                    "PaymentMethod": getcol("Method").astype(str) if mappings["Method"].get() else None,
                })
                # Fill Type if missing: sign-based heuristic
                nd["Type"] = nd["Type"].where(nd["Type"].isin(["income","expense"]))
                nd.loc[nd["Type"].isna() & (nd["Amount"] < 0), "Type"] = "expense"
                nd.loc[nd["Type"].isna() & (nd["Amount"] > 0), "Type"] = "income"
                nd["Type"] = nd["Type"].fillna(def_type.get())
                # Normalize Amount to positive numbers
                nd.loc[nd["Type"]=="expense","Amount"] = nd.loc[nd["Type"]=="expense","Amount"].abs()
                nd.loc[nd["Type"]=="income","Amount"] = nd.loc[nd["Type"]=="income","Amount"].abs()
                # Category default
                nd["Category"] = nd["Category"].where(nd["Category"].notna() & (nd["Category"].astype(str)!="nan"), def_cat.get())
                # Required columns for sheet
                nd["Reference"] = ""
                nd["LinkedDoc"] = ""
                # Drop unusable rows
                nd = nd.dropna(subset=["Date","Amount"])
                # Append and write
                df = pd.concat([df, nd], ignore_index=True)
                write_sheet(df, "Transactions")
                self._toast(f"Imported {len(nd)} rows")
                win.destroy()
                self._refresh_tx_table()
                self._load_dashboard()
            except Exception as e:
                messagebox.showerror("Import failed", str(e))
        ttk.Button(btns, text="Import", command=do_import).pack(side="right", padx=6)

    # ----- Invoices -----
    def _build_invoices_tab(self):
        tab = ttk.Frame(self.nb); self.nb.add(tab, text="Invoices")
        form = self._card(tab); form.pack(fill="x", padx=6, pady=6)
        r1 = ttk.Frame(form); r1.pack(fill="x", pady=4)
        ttk.Label(r1, text="Date").pack(side="left"); 
        self.inv_date = ttk.Entry(r1, width=12); self.inv_date.insert(0, datetime.today().strftime("%Y-%m-%d")); self.inv_date.pack(side="left", padx=6)
        ttk.Label(r1, text="Due Date").pack(side="left", padx=(12,6))
        self.inv_due = ttk.Entry(r1, width=12); self.inv_due.insert(0, (datetime.today()+timedelta(days=14)).strftime("%Y-%m-%d")); self.inv_due.pack(side="left")
        ttk.Label(r1, text="Customer").pack(side="left", padx=(12,6))
        self.inv_cust = ttk.Combobox(r1, values=get_customers(), width=28); self.inv_cust.pack(side="left", padx=6)

        r2 = ttk.Frame(form); r2.pack(fill="x", pady=4)
        ttk.Label(r2, text="Item").pack(side="left"); self.inv_item = ttk.Entry(r2, width=28); self.inv_item.pack(side="left", padx=6)
        ttk.Label(r2, text="Qty").pack(side="left", padx=(12,6)); self.inv_qty = ttk.Entry(r2, width=8); self.inv_qty.insert(0,"1"); self.inv_qty.pack(side="left")
        ttk.Label(r2, text="Rate").pack(side="left", padx=(12,6)); self.inv_rate = ttk.Entry(r2, width=12); self.inv_rate.insert(0,"0.00"); self.inv_rate.pack(side="left")
        ttk.Label(r2, text="Notes").pack(side="left", padx=(12,6)); self.inv_notes = ttk.Entry(r2, width=36); self.inv_notes.pack(side="left", padx=6)

        r3 = ttk.Frame(form); r3.pack(fill="x", pady=4)
        FancyButton(r3, text="Create Invoice", command=self._add_invoice).pack(side="left", padx=4)
        ttk.Label(r3, text="Mark Paid Invoice ID").pack(side="left", padx=(18,6))
        self.inv_mark_id = ttk.Entry(r3, width=12); self.inv_mark_id.pack(side="left", padx=4)
        ttk.Button(r3, text="Mark Paid", command=self._mark_invoice_paid).pack(side="left", padx=4)

        self.inv_table = ttk.Treeview(tab, columns=("InvoiceID","Date","DueDate","Customer","Item","Qty","Rate","Amount","Status","Notes"), show="headings", height=16)
        for c,w in [("InvoiceID",100),("Date",100),("DueDate",100),("Customer",180),("Item",220),("Qty",60),("Rate",90),("Amount",110),("Status",90),("Notes",220)]:
            self.inv_table.heading(c, text=c); self.inv_table.column(c, width=w, anchor="center" if c in ("Qty","Rate","Amount","Status") else "w")
        self.inv_table.pack(fill="both", expand=True, padx=6, pady=(0,6))
        self._refresh_inv_table()

    def _add_invoice(self):
        try:
            inv_id, amt = create_invoice(
                self.inv_date.get(), self.inv_due.get(), self.inv_cust.get(),
                self.inv_item.get(), float(self.inv_qty.get()), float(self.inv_rate.get()),
                self.inv_notes.get()
            )
            self._toast(f"Invoice {inv_id} created for ${amt:,.2f}")
            self._refresh_inv_table()
            self._refresh_tx_table()
            self._load_dashboard()
            self._confetti()
        except Exception as e:
            messagebox.showerror("Error", f"Create invoice failed:\n{e}")

    def _mark_invoice_paid(self):
        try:
            iid = self.inv_mark_id.get().strip()
            if not iid:
                messagebox.showwarning("Missing", "Enter an Invoice ID")
                return
            mark_invoice_paid(iid, datetime.today().strftime("%Y-%m-%d"))
            self._toast(f"{iid} marked paid")
            self._refresh_inv_table()
            self._refresh_tx_table()
            self._load_dashboard()
        except Exception as e:
            messagebox.showerror("Error", f"Mark paid failed:\n{e}")

    def _refresh_inv_table(self):
        for i in self.inv_table.get_children():
            self.inv_table.delete(i)
        try:
            df = read_sheet("Invoices")
            if df.empty: return
            df = df.copy()
            for col in ("Date","DueDate"):
                if col in df.columns:
                    df[col] = pd.to_datetime(df[col]).dt.strftime("%Y-%m-%d")
            for _, r in df.iterrows():
                self.inv_table.insert("", "end", values=(
                    r.get("InvoiceID",""), r.get("Date",""), r.get("DueDate",""),
                    r.get("CustomerName",""), r.get("Item",""), r.get("Qty",""),
                    f"{float(r.get('Rate',0.0)):.2f}", f"${float(r.get('Amount',0.0)):,.2f}",
                    r.get("Status",""), r.get("Notes","")
                ))
        except Exception as e:
            messagebox.showerror("Error", f"Load invoices failed:\n{e}")

    # ----- Customers -----
    def _build_customers_tab(self):
        tab = ttk.Frame(self.nb); self.nb.add(tab, text="Customers")
        form = self._card(tab); form.pack(fill="x", padx=6, pady=6)
        r1 = ttk.Frame(form); r1.pack(fill="x", pady=4)
        ttk.Label(r1, text="Name").pack(side="left"); self.cu_name = ttk.Entry(r1, width=28); self.cu_name.pack(side="left", padx=6)
        ttk.Label(r1, text="Email").pack(side="left", padx=(12,6)); self.cu_email = ttk.Entry(r1, width=28); self.cu_email.pack(side="left", padx=6)
        r2 = ttk.Frame(form); r2.pack(fill="x", pady=4)
        ttk.Label(r2, text="Phone").pack(side="left"); self.cu_phone = ttk.Entry(r2, width=20); self.cu_phone.pack(side="left", padx=6)
        ttk.Label(r2, text="Billing Address").pack(side="left", padx=(12,6)); self.cu_addr = ttk.Entry(r2, width=50); self.cu_addr.pack(side="left", padx=6)
        r3 = ttk.Frame(form); r3.pack(fill="x", pady=4)
        ttk.Label(r3, text="Notes").pack(side="left"); self.cu_notes = ttk.Entry(r3, width=60); self.cu_notes.pack(side="left", padx=6)
        FancyButton(form, text="Add Customer", command=self._add_customer).pack(pady=6)

        self.cu_table = ttk.Treeview(tab, columns=("CustomerName","Email","Phone","BillingAddress","Notes"), show="headings", height=16)
        for c,w in [("CustomerName",200),("Email",220),("Phone",130),("BillingAddress",320),("Notes",240)]:
            self.cu_table.heading(c, text=c); self.cu_table.column(c, width=w, anchor="w")
        self.cu_table.pack(fill="both", expand=True, padx=6, pady=(0,6))
        self._refresh_customers_table()

    def _add_customer(self):
        try:
            add_customer(self.cu_name.get(), self.cu_email.get(), self.cu_phone.get(), self.cu_addr.get(), self.cu_notes.get())
            self._toast("Customer added")
            self.inv_cust.configure(values=get_customers())
            self._refresh_customers_table()
        except Exception as e:
            messagebox.showerror("Error", f"Add customer failed:\n{e}")

    def _refresh_customers_table(self):
        for i in self.cu_table.get_children():
            self.cu_table.delete(i)
        try:
            df = read_sheet("Customers")
            if df.empty: return
            for _, r in df.iterrows():
                self.cu_table.insert("", "end", values=(r.get("CustomerName",""), r.get("Email",""), r.get("Phone",""), r.get("BillingAddress",""), r.get("Notes","")))
        except Exception as e:
            messagebox.showerror("Error", f"Load customers failed:\n{e}")

    # ----- Payroll -----
    def _build_payroll_tab(self):
        tab = ttk.Frame(self.nb); self.nb.add(tab, text="Payroll")
        form = self._card(tab); form.pack(fill="x", padx=6, pady=6)
        r1 = ttk.Frame(form); r1.pack(fill="x", pady=4)
        ttk.Label(r1, text="Employee Name").pack(side="left"); self.pay_empname = ttk.Entry(r1, width=24); self.pay_empname.pack(side="left", padx=6)
        ttk.Label(r1, text="Type").pack(side="left", padx=(12,6))
        self.pay_type = ttk.Combobox(r1, values=["hourly","salary"], state="readonly", width=10); self.pay_type.current(0); self.pay_type.pack(side="left")
        ttk.Label(r1, text="Hourly Rate").pack(side="left", padx=(12,6)); self.pay_hrate = ttk.Entry(r1, width=10); self.pay_hrate.insert(0,"0.00"); self.pay_hrate.pack(side="left")
        ttk.Label(r1, text="Salary").pack(side="left", padx=(12,6)); self.pay_salary = ttk.Entry(r1, width=12); self.pay_salary.insert(0,"0.00"); self.pay_salary.pack(side="left")
        ttk.Label(r1, text="Tax Rate").pack(side="left", padx=(12,6)); self.pay_tax = ttk.Entry(r1, width=8); self.pay_tax.insert(0,"0.10"); self.pay_tax.pack(side="left")
        FancyButton(form, text="Add Employee", command=self._add_employee).pack(pady=6)

        run = self._card(tab); run.pack(fill="x", padx=6, pady=6)
        r2 = ttk.Frame(run); r2.pack(fill="x", pady=4)
        ttk.Label(r2, text="Date").pack(side="left"); self.run_date = ttk.Entry(r2, width=12); self.run_date.insert(0, datetime.today().strftime("%Y-%m-%d")); self.run_date.pack(side="left", padx=6)
        ttk.Label(r2, text="Employee").pack(side="left", padx=(12,6)); self.run_emp = ttk.Combobox(r2, values=get_employees(), width=28); self.run_emp.pack(side="left")
        ttk.Label(r2, text="Hours (hourly)").pack(side="left", padx=(12,6)); self.run_hours = ttk.Entry(r2, width=8); self.run_hours.insert(0,"0"); self.run_hours.pack(side="left")
        FancyButton(run, text="Run Payroll", command=self._run_payroll).pack(pady=6)

        self.pay_table = ttk.Treeview(tab, columns=("PayslipID","Date","Employee","Hours","Gross","Tax","Net","Notes"), show="headings", height=14)
        for c,w in [("PayslipID",100),("Date",100),("Employee",180),("Hours",80),("Gross",110),("Tax",110),("Net",110),("Notes",220)]:
            self.pay_table.heading(c, text=c); self.pay_table.column(c, width=w, anchor="center" if c in ("Hours","Date") else "w")
        self.pay_table.pack(fill="both", expand=True, padx=6, pady=(0,6))
        self._refresh_pay_table()

    def _add_employee(self):
        try:
            add_employee(
                self.pay_empname.get(), self.pay_type.get(),
                float(self.pay_hrate.get() or 0.0), float(self.pay_salary.get() or 0.0),
                float(self.pay_tax.get() or 0.1), ""
            )
            self._toast("Employee added")
            self.run_emp.configure(values=get_employees())
        except Exception as e:
            messagebox.showerror("Error", f"Add employee failed:\n{e}")

    def _run_payroll(self):
        try:
            ps_id, gross, tax, net = run_payroll(self.run_date.get(), self.run_emp.get(), float(self.run_hours.get() or 0.0))
            self._toast(f"Payroll {ps_id}: Net ${net:,.2f}")
            self._refresh_pay_table()
            self._refresh_tx_table()
            self._load_dashboard()
        except Exception as e:
            messagebox.showerror("Error", f"Run payroll failed:\n{e}")

    def _refresh_pay_table(self):
        for i in self.pay_table.get_children():
            self.pay_table.delete(i)
        try:
            df = read_sheet("Payslips")
            if df.empty: return
            df = df.copy()
            df["Date"] = pd.to_datetime(df["Date"]).dt.strftime("%Y-%m-%d")
            for _, r in df.iterrows():
                self.pay_table.insert("", "end", values=(
                    r.get("PayslipID",""), r.get("Date",""), r.get("EmployeeName",""),
                    r.get("Hours",""), f"${float(r.get('Gross',0.0)):,.2f}",
                    f"${float(r.get('Tax',0.0)):,.2f}", f"${float(r.get('Net',0.0)):,.2f}",
                    r.get("Notes","")
                ))
        except Exception as e:
            messagebox.showerror("Error", f"Load payslips failed:\n{e}")

    # ----- Reports -----
    def _build_reports_tab(self):
        tab = ttk.Frame(self.nb); self.nb.add(tab, text="Reports")
        top = ttk.Frame(tab); top.pack(fill="x", pady=6)
        FancyButton(top, text="Build / Refresh Reports", command=self._load_report_preview).pack(side="left", padx=6)
        self.rep_table = ttk.Treeview(tab, columns=("Col1","Col2","Col3","Col4","Col5"), show="headings", height=20)
        for i in range(1,6):
            self.rep_table.heading(f"Col{i}", text=f"Col{i}")
            self.rep_table.column(f"Col{i}", width=160, anchor="w")
        self.rep_table.pack(fill="both", expand=True, padx=6, pady=6)
        self._load_report_preview()

    def _load_report_preview(self):
        try:
            build_reports()  # writes to sheet
            df = pd.read_excel(EXCEL_PATH, sheet_name="Reports", header=0)
            # render any table-ish content – keep it simple
            for i in self.rep_table.get_children():
                self.rep_table.delete(i)
            # craft display: take up to 5 columns
            df = df.fillna("")
            max_cols = min(5, df.shape[1])
            self.rep_table["columns"] = [f"Col{i}" for i in range(1,max_cols+1)]
            for i,c in enumerate(df.columns[:max_cols], 1):
                self.rep_table.heading(f"Col{i}", text=c)
            for _, r in df.iloc[:500].iterrows():
                vals = [str(r[c]) for c in df.columns[:max_cols]]
                self.rep_table.insert("", "end", values=vals)
        except Exception as e:
            messagebox.showerror("Error", f"Build/preview failed:\n{e}")

    # ----- Settings -----
    def _build_settings_tab(self):
        tab = ttk.Frame(self.nb); self.nb.add(tab, text="Settings")
        card = self._card(tab); card.pack(fill="x", padx=6, pady=6)
        ttk.Label(card, text="Company Name").pack(side="left"); 
        self.set_company = ttk.Entry(card, width=40); self.set_company.insert(0, get_company_name()); self.set_company.pack(side="left", padx=6)
        ttk.Button(card, text="Save", command=self._save_company).pack(side="left", padx=6)

        cat = self._card(tab); cat.pack(fill="both", expand=True, padx=6, pady=6)
        ttk.Label(cat, text="Categories", font=("Segoe UI", 11, "bold")).pack(anchor="w")
        row = ttk.Frame(cat); row.pack(fill="x", pady=4)
        ttk.Label(row, text="Name").pack(side="left"); self.cat_name = ttk.Entry(row, width=28); self.cat_name.pack(side="left", padx=6)
        ttk.Label(row, text="Type").pack(side="left", padx=(12,6))
        self.cat_type = ttk.Combobox(row, values=["income","expense"], state="readonly", width=12); self.cat_type.current(1); self.cat_type.pack(side="left")
        ttk.Button(row, text="Add", command=self._add_category).pack(side="left", padx=6)
        ttk.Button(row, text="Remove Selected", command=self._remove_category).pack(side="left", padx=6)

        self.cat_table = ttk.Treeview(cat, columns=("Category","Type"), show="headings", height=14)
        for c,w in [("Category",320),("Type",120)]:
            self.cat_table.heading(c, text=c); self.cat_table.column(c, width=w, anchor="w")
        self.cat_table.pack(fill="both", expand=True, pady=6)
        self._refresh_categories_table()

    def _save_company(self):
        try:
            set_company_name(self.set_company.get())
            self._toast("Saved")
            self.after(50, lambda: self.banner.draw_gradient())
        except Exception as e:
            messagebox.showerror("Error", f"Save failed:\n{e}")

    def _add_category(self):
        try:
            add_category(self.cat_name.get(), self.cat_type.get())
            self._refresh_categories_table()
            self.tx_cat.configure(values=get_categories())
            self._toast("Category added")
        except Exception as e:
            messagebox.showerror("Error", f"Add category failed:\n{e}")

    def _remove_category(self):
        try:
            sel = self.cat_table.selection()
            if not sel: return
            name = self.cat_table.item(sel[0], "values")[0]
            remove_category(name)
            self._refresh_categories_table()
            self.tx_cat.configure(values=get_categories())
            self._toast("Category removed")
        except Exception as e:
            messagebox.showerror("Error", f"Remove failed:\n{e}")

    def _refresh_categories_table(self):
        for i in self.cat_table.get_children():
            self.cat_table.delete(i)
        try:
            df = get_categories_df()
            for _, r in df.iterrows():
                self.cat_table.insert("", "end", values=(r.get("Category",""), r.get("Type","")))
        except Exception as e:
            messagebox.showerror("Error", f"Load categories failed:\n{e}")

    # ----- Help -----
    def _build_help_tab(self):
        tab = ttk.Frame(self.nb); self.nb.add(tab, text="Help")
        top = ttk.Frame(tab); top.pack(fill="x", padx=8, pady=8)
        ttk.Label(top, text="Search").pack(side="left")
        q = ttk.Entry(top, width=40); q.pack(side="left", padx=6)
        lst = tk.Listbox(tab, height=18); lst.pack(fill="both", expand=True, padx=8, pady=8)
        body = tk.Text(tab, height=10, wrap="word"); body.pack(fill="both", expand=False, padx=8, pady=(0,8))
        kb = HELP_CONTENT[:]

        def refresh_list():
            lst.delete(0, tk.END)
            term = q.get().lower().strip()
            for item in kb:
                label = f"{item['topic']} — {item['q']}"
                if term in label.lower() or term in item["a"].lower():
                    lst.insert(tk.END, label)
        def show_item(_=None):
            idx = lst.curselection()
            if not idx: return
            item = [k for k in kb if f"{k['topic']} — {k['q']}" == lst.get(idx)][0]
            body.delete("1.0", tk.END)
            body.insert("1.0", f"{item['q']}\n\n{item['a']}")

        q.bind("<KeyRelease>", lambda e: refresh_list())
        lst.bind("<<ListboxSelect>>", show_item)
        refresh_list()
        if lst.size(): lst.selection_set(0); show_item()

    # ----- Command Palette -----
    def _register_commands(self):
        cmds = [
            ("Dashboard: Refresh", self._load_dashboard),
            ("Transactions: Add", self._add_tx),
            ("Transactions: Import CSV", self._import_csv_wizard),
            ("Transactions: Export CSV", self._export_transactions_csv),
            ("Invoices: Create", self._add_invoice),
            ("Invoices: Mark Paid", self._mark_invoice_paid),
            ("Reports: Build/Refresh", self._load_report_preview),
            ("Settings: Save Company", self._save_company),
            ("Open: Workbook", self._open_workbook),
            ("Open: Month Folder", self._open_month_folder),
            ("Month: Switch…", self._switch_month_dialog),
            ("Month: Close (Finalize)", self._close_month),
            ("Backup: Workbook", self._backup_workbook),
        ]
        self._commands = cmds

    def _open_command_palette(self):
        if not hasattr(self, "_commands"): self._register_commands()
        CommandPalette(self, self._commands)

# ---- main ----
if __name__ == "__main__":
    try:
        app = RainbowLedgerApp()
        app.mainloop()
    except Exception as e:
        messagebox.showerror("Fatal Error", str(e))
